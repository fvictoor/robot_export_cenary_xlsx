import os
import argparse
from collections import Counter
from openpyxl import Workbook
from robot.api import TestSuiteBuilder

def exportar_testes_para_excel(pasta_testes, caminho_arquivo_excel):
    """
    Varre uma pasta em busca de arquivos .robot, extrai informações dos cenários
    de teste e exporta os dados para um arquivo Excel com colunas de tag estruturadas e ordenadas.

    Args:
        pasta_testes (str): O caminho para a pasta contendo os arquivos .robot.
        caminho_arquivo_excel (str): O caminho completo onde o arquivo Excel será salvo.
    """
    TIPOS_DE_TESTE = ['frontend', 'api']
    PRIORIDADES = ['alta', 'media', 'baixa']

    dados_testes_estruturados = []
    contador_tags_geral = Counter()
    resumo_por_arquivo = {}
    max_outras_tags = 0
    total_testes = 0
    arquivos_processados = 0

    print(f"\n🔍 Analisando testes em: {pasta_testes}")

    if not os.path.isdir(pasta_testes):
        print(f"❌ Erro: O diretório de entrada '{pasta_testes}' não foi encontrado ou não é um diretório.")
        return
    
    # Define o caminho base para tornar os caminhos dos arquivos relativos
    caminho_base_abs = os.path.abspath(pasta_testes)

    for root, dirs, files in os.walk(caminho_base_abs):
        for file in files:
            if file.endswith(".robot"):
                caminho_completo = os.path.join(root, file)
                try:
                    suite = TestSuiteBuilder().build(caminho_completo)
                    arquivos_processados += 1
                    qtd_testes_arquivo = 0

                    for test in suite.tests:
                        if not test.name: continue
                        
                        total_testes += 1
                        qtd_testes_arquivo += 1
                        contador_tags_geral.update(test.tags)

                        modulo = ""
                        tipo_teste = ""
                        prioridade = ""
                        
                        tags_nao_classificadas = []

                        for tag in test.tags:
                            tag_lower = tag.lower()
                            if tag_lower in PRIORIDADES:
                                prioridade = tag
                            elif tag_lower in TIPOS_DE_TESTE:
                                tipo_teste = tag
                            else:
                                tags_nao_classificadas.append(tag)
                        
                        if tags_nao_classificadas:
                            modulo = tags_nao_classificadas.pop(0)
                        
                        outras_tags = tags_nao_classificadas
                        max_outras_tags = max(max_outras_tags, len(outras_tags))

                        # --- ALTERAÇÃO APLICADA AQUI ---
                        # Gera o caminho relativo do arquivo a partir do diretório de entrada
                        caminho_relativo = os.path.relpath(caminho_completo, os.path.dirname(caminho_base_abs))

                        dados_testes_estruturados.append({
                            "Arquivo": caminho_relativo, # Utiliza o novo caminho relativo
                            "Nome do Teste": test.name,
                            "Documentação": test.doc,
                            "Módulo": modulo,
                            "Tipo de Teste": tipo_teste,
                            "Prioridade": prioridade,
                            "Outras Tags": outras_tags
                        })

                    if qtd_testes_arquivo > 0:
                        # Usamos o caminho relativo para o resumo também
                        resumo_path_relativo = os.path.relpath(caminho_completo, os.path.dirname(caminho_base_abs))
                        resumo_por_arquivo[resumo_path_relativo] = qtd_testes_arquivo

                except Exception as e:
                    print(f"⚠️  Aviso: Erro ao processar o arquivo {file}: {e}")

    if total_testes == 0:
        print("🔴 Nenhum cenário de teste foi encontrado nos arquivos .robot.")
        return

    dados_testes_estruturados.sort(key=lambda x: (str(x.get('Módulo', '')), str(x.get('Nome do Teste', ''))))
        
    wb = Workbook()

    ws = wb.active
    ws.title = "Cenários de Testes"
    cabecalhos_base = ["Arquivo", "Nome do Teste", "Documentação", "Módulo", "Tipo de Teste", "Prioridade"]
    cabecalhos_extras = [f"Tag Extra {i+1}" for i in range(max_outras_tags)]
    ws.append(cabecalhos_base + cabecalhos_extras)

    for item in dados_testes_estruturados:
        linha = [
            item["Arquivo"],
            item["Nome do Teste"],
            item["Documentação"],
            item["Módulo"],
            item["Tipo de Teste"],
            item["Prioridade"],
            *item["Outras Tags"]
        ]
        ws.append(linha)

    ws_resumo = wb.create_sheet(title="Resumo")
    ws_resumo.append(["Arquivo", "Quantidade de Testes"])
    for arquivo, quantidade in sorted(resumo_por_arquivo.items()):
        ws_resumo.append([arquivo, quantidade])
    ws_resumo.append(["TOTAL", total_testes])

    ws_tags = wb.create_sheet(title="Tags")
    ws_tags.append(["Tag", "Quantidade"])
    for tag, count in contador_tags_geral.most_common():
        ws_tags.append([tag, count])

    try:
        output_dir = os.path.dirname(caminho_arquivo_excel)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            
        wb.save(caminho_arquivo_excel)

        print("\n--- RESUMO DA EXECUÇÃO ---")
        print(f"📦 Arquivos .robot processados: {arquivos_processados}")
        print(f"✅ Total de cenários encontrados: {total_testes}")
        print(f"🏷️  Tags únicas encontradas: {len(contador_tags_geral)}")
        print(f"💾 Relatório Excel salvo em: {caminho_arquivo_excel}")

    except Exception as e:
        print(f"❌ Erro ao salvar o arquivo Excel: {e}")


def main():
    parser = argparse.ArgumentParser(
        description="Exporta cenários de teste do Robot Framework para uma planilha Excel estruturada.",
        formatter_class=argparse.HelpFormatter
    )
    parser.add_argument(
        '--testinput',
        type=str,
        required=True,
        help='Caminho para o diretório raiz contendo os arquivos de teste .robot.'
    )
    parser.add_argument(
        '--outputdir',
        type=str,
        default='.',
        help='Diretório onde o arquivo Excel será salvo. Padrão: diretório atual.'
    )
    args = parser.parse_args()

    input_folder_name = os.path.basename(os.path.normpath(args.testinput))
    output_filename = f"cenarios_{input_folder_name}.xlsx"
    output_filepath = os.path.join(args.outputdir, output_filename)

    exportar_testes_para_excel(args.testinput, output_filepath)

if __name__ == "__main__":
    main()